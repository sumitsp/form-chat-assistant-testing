from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
import json
import re
from typing import Any

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from backend.eligibility import EligibilityRequest, find_eligible_programs
from backend.loanpass_client import list_program_products
from backend.pdf import (
    ProfileRow,
    ProfileSection,
    ScenarioPdfProgramItem,
    ScenarioPdfRejectedItem,
    ScenarioPdfRequest,
    generate_scenario_pdf_bytes,
    humanize_reject_reason,
    rejected_programs_from_trace,
)


_MEDIA_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_MEDIA_PDF = "application/pdf"
_SCENARIO_NAME_FIELD = "scenarioName"


def excel_media_type() -> str:
    return _MEDIA_XLSX


def pdf_media_type() -> str:
    return _MEDIA_PDF


def batch_frontend_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>Mortgage Batch Runner</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 0; padding: 24px; background: #f8fafc; color: #0f172a; }
    .wrap { max-width: 760px; margin: 0 auto; background: white; border: 1px solid #e2e8f0; border-radius: 12px; padding: 20px; }
    h1 { margin: 0 0 8px; font-size: 24px; }
    p { margin: 8px 0; line-height: 1.45; }
    .row { margin-top: 16px; display: flex; gap: 10px; flex-wrap: wrap; }
    button, .linkbtn {
      border: 1px solid #cbd5e1; background: #0b3b71; color: white; border-radius: 8px; padding: 10px 14px;
      font-size: 14px; cursor: pointer; text-decoration: none; display: inline-block;
    }
    .linkbtn { background: #334155; }
    input[type=file] { margin-top: 6px; width: 100%; }
    #status { margin-top: 14px; color: #334155; }
    .help { margin-top: 14px; font-size: 13px; color: #475569; }
    code { background: #f1f5f9; padding: 2px 4px; border-radius: 4px; }
    .busy {
      opacity: 0.6;
      filter: grayscale(0.25);
      pointer-events: none;
      user-select: none;
    }
    button:disabled, input:disabled, .linkbtn[aria-disabled="true"] {
      cursor: not-allowed;
      opacity: 0.75;
    }
  </style>
</head>
<body>
  <div class="wrap" id="panel">
    <h1>Mortgage Batch Runner</h1>
    <p>Upload one Excel file where each row is one scenario. The server runs eligibility + LoanPASS checks and returns a PDF report.</p>
    <div class="row">
      <a class="linkbtn" id="downloadTemplateBtn" href="/api/batch/template">Download Excel Template</a>
    </div>
    <form id="uploadForm">
      <p><strong>Select Excel:</strong></p>
      <input id="xlsx" type="file" accept=".xlsx" required />
      <div class="row">
        <button id="runBtn" type="submit">Run and Download PDF</button>
      </div>
    </form>
    <div class="row" id="downloadsRow" style="display:none;">
      <a class="linkbtn" id="downloadPdfBtn" href="#">Download PDF</a>
      <a class="linkbtn" id="downloadJsonBtn" href="#">Download JSON</a>
    </div>
    <div id="status"></div>
    <p class="help">Template headers use API keys (for example <code>occupancy</code>, <code>loanAmount</code>, <code>state</code>, <code>documentationType</code>). Common form-question labels are also accepted.</p>
  </div>
  <script>
    const panel = document.getElementById("panel");
    const form = document.getElementById("uploadForm");
    const input = document.getElementById("xlsx");
    const runBtn = document.getElementById("runBtn");
    const downloadTemplateBtn = document.getElementById("downloadTemplateBtn");
    const downloadsRow = document.getElementById("downloadsRow");
    const downloadPdfBtn = document.getElementById("downloadPdfBtn");
    const downloadJsonBtn = document.getElementById("downloadJsonBtn");
    const status = document.getElementById("status");
    let inFlight = false;
    let requestController = null;
    let requestTimeoutId = null;

    function setBusy(isBusy) {
      inFlight = isBusy;
      panel.classList.toggle("busy", isBusy);
      runBtn.disabled = isBusy;
      input.disabled = isBusy;
      if (isBusy) {
        downloadTemplateBtn.setAttribute("aria-disabled", "true");
        downloadTemplateBtn.setAttribute("tabindex", "-1");
      } else {
        downloadTemplateBtn.removeAttribute("aria-disabled");
        downloadTemplateBtn.removeAttribute("tabindex");
      }
    }

    downloadTemplateBtn.addEventListener("click", (e) => {
      if (inFlight) e.preventDefault();
    });

    form.addEventListener("submit", async (e) => {
      e.preventDefault();
      if (inFlight) return;
      const file = input.files?.[0];
      if (!file) {
        status.textContent = "Please choose an .xlsx file.";
        return;
      }
      requestController = new AbortController();
      requestTimeoutId = window.setTimeout(() => {
        if (requestController) requestController.abort();
      }, 120000);
      setBusy(true);
      downloadsRow.style.display = "none";
      status.textContent = "Running scenarios... this may take a moment.";
      const fd = new FormData();
      fd.append("file", file);
      try {
        const res = await fetch("/api/batch/run", {
          method: "POST",
          body: fd,
          signal: requestController.signal,
        });
        if (!res.ok) {
          const txt = await res.text();
          throw new Error(txt || ("Request failed with status " + res.status));
        }
        const data = await res.json();
        downloadPdfBtn.href = data.pdf_download_url;
        downloadPdfBtn.download = "mortgage-batch-report.pdf";
        downloadJsonBtn.href = data.json_download_url;
        downloadJsonBtn.download = "eligibility_batch_test_api_contract.json";
        downloadsRow.style.display = "flex";
        status.textContent = "Done. Use the download buttons.";
      } catch (err) {
        if (err?.name === "AbortError") {
          status.textContent = "Timed out or cancelled. Please try again.";
          return;
        }
        status.textContent = "Failed: " + (err?.message || err);
      } finally {
        if (requestTimeoutId) {
          window.clearTimeout(requestTimeoutId);
          requestTimeoutId = null;
        }
        requestController = null;
        setBusy(false);
      }
    });

    window.addEventListener("beforeunload", () => {
      if (requestController) requestController.abort();
    });

    // Refresh/new load always starts from home state.
    status.textContent = "";
    setBusy(false);
  </script>
</body>
</html>
"""


_ELIGIBILITY_KEYS = list(EligibilityRequest.model_fields.keys())
_LOWER_KEY_MAP = {k.lower(): k for k in _ELIGIBILITY_KEYS}

_OPTIONAL_FIELDS = {
    "existingFirstLien",
    "cltv",
    "dscr",
    "creditEvent",
    "creditEventType",
    "yearsSinceEvent",
    "firstTimeHomebuyer",
    "rentalType",
    "qualificationPath",
    "firstTimeInvestor",
    "establishedPrimaryRes",
    "stateCounty",
    "stateCity",
    "stateBorough",
    "stateZipCode",
    "isInBaltimoreCity",
    "isInIndianapolis",
    "isInPhiladelphia",
    "isInMemphis",
    "isInLubbock",
    "loanTerm",
    "interestOnlyPref",
    "rateTypePref",
    "secondLienProduct",
    "hiLavaZone",
    "isRuralProperty",
    "acreage",
    "nonOccupantCoBorrower",
    "combinedDti",
    "visaType",
    "visaCategory",
    "hasUsCredit",
    "prepayStepdown",
    "listingSeasoning",
}

_DROPDOWN_OPTIONS: dict[str, list[str]] = {
    "occupancy": ["Primary Residence", "Second Home", "Investment Property"],
    "loanPurpose": ["Purchase", "Refinance", "Cash-Out Refinance"],
    "primaryLoanPurpose": ["Purchase", "Refinance", "Cash-Out Refinance"],
    "documentationType": [
        "Full Documentation",
        "Bank Statements (12 or 24 Months)",
        "1099",
        "Asset Utilization",
        "P&L with 2 month Bank Statement",
        "Alternative Documentation",
        "Rental Income",
    ],
    "propertyType": [
        "single_family",
        "pud",
        "townhouse",
        "condo_warrantable",
        "condo_non_warrantable",
        "condotel",
        "two_to_four_family",
        "five_to_eight_unit",
        "mixed_use",
        "manufactured_home",
        "cooperative",
    ],
    "citizenship": [
        "US Citizen",
        "Permanent Resident Alien",
        "Non-Permanent Resident Alien",
        "Foreign National",
    ],
    "paymentHistory": ["0x30x12", "1x30x12", "0x60x12", "1x60x12"],
    "lienPosition": ["first_lien_only", "second_lien", "second_lien_piggyback"],
    "secondLienProduct": ["HELOC", "HELOAN"],
    "investmentIncomePath": ["income", "dscr"],
    "qualificationPath": ["DTI", "DSCR"],
    "rentalType": ["Long-term", "Short-term"],
    "loanTerm": ["30", "40", "No Preference"],
    "interestOnlyPref": ["Yes", "No", "No Preference"],
    "rateTypePref": ["Fixed", "ARM", "No Preference"],
    "visaCategory": ["A", "E", "G", "H", "L", "O", "TN", "Other"],
    "prepayStepdown": ["No", "5-4-3-2-1", "3-2-1"],
    "prepaymentTerms": ["No", "1", "2", "3", "4", "5"],
    "firstTimeHomebuyer": ["Yes", "No"],
    "firstTimeInvestor": ["Yes", "No"],
    "isSecondLien": ["Yes", "No"],
    "hiLavaZone": ["Yes", "No"],
    "isRuralProperty": ["Yes", "No"],
    "nonOccupantCoBorrower": ["Yes", "No"],
    "ofacSanctioned": ["Yes", "No"],
    "hasUsCredit": ["Yes", "No"],
    "listingSeasoning": ["Yes", "No"],
    "powerOfAttorney": ["Yes", "No"],
    "nonArmsLength": ["Yes", "No"],
}


_HEADER_ALIASES = {
    "occupancy": "occupancy",
    "loan purpose": "loanPurpose",
    "primary loan purpose": "primaryLoanPurpose",
    "state": "state",
    "property value": "valueSalesPrice",
    "value/sales price": "valueSalesPrice",
    "sales price": "valueSalesPrice",
    "loan amount": "loanAmount",
    "ltv": "ltv",
    "estimated dti": "estimatedDti",
    "documentation type": "documentationType",
    "prepayment terms": "prepaymentTerms",
    "property type": "propertyType",
    "citizenship": "citizenship",
    "decision credit score": "decisionCreditScore",
    "existing first lien": "existingFirstLien",
    "cltv": "cltv",
    "dscr": "dscr",
    "credit event": "creditEvent",
    "credit event type": "creditEventType",
    "years since event": "yearsSinceEvent",
    "first time homebuyer": "firstTimeHomebuyer",
    "rental type": "rentalType",
    "qualification path": "qualificationPath",
    "payment history": "paymentHistory",
    "first time investor": "firstTimeInvestor",
    "established primary residence": "establishedPrimaryRes",
    "is second lien": "isSecondLien",
    "state county": "stateCounty",
    "state city": "stateCity",
    "state borough": "stateBorough",
    "state zip code": "stateZipCode",
    "is in baltimore city": "isInBaltimoreCity",
    "is in indianapolis": "isInIndianapolis",
    "is in philadelphia": "isInPhiladelphia",
    "is in memphis": "isInMemphis",
    "is in lubbock": "isInLubbock",
    "loan term": "loanTerm",
    "interest only pref": "interestOnlyPref",
    "rate type pref": "rateTypePref",
    "lien position": "lienPosition",
    "second lien product": "secondLienProduct",
    "hi lava zone": "hiLavaZone",
    "is rural property": "isRuralProperty",
    "acreage": "acreage",
    "non occupant co borrower": "nonOccupantCoBorrower",
    "combined dti": "combinedDti",
    "visa type": "visaType",
    "visa category": "visaCategory",
    "ofac sanctioned": "ofacSanctioned",
    "has us credit": "hasUsCredit",
    "investment income path": "investmentIncomePath",
    "prepay stepdown": "prepayStepdown",
    "listing seasoning": "listingSeasoning",
    "power of attorney": "powerOfAttorney",
    "non arms length": "nonArmsLength",
}


def _normalize_header(text: str) -> str:
    s = (text or "").strip().lower()
    s = re.sub(r"\s*\(\s*optional\s*\)\s*", " ", s)
    s = re.sub(r"\[\s*dropdown\s*\]", " ", s)
    s = re.sub(r"\(\s*dropdown\s*\)", " ", s)
    s = re.sub(r"\(\s*select\s*\)", " ", s)
    s = s.replace("▼", " ")
    s = s.replace("*", " ")
    s = re.sub(r"[_/\\-]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _header_to_field(header: str) -> str | None:
    raw = (header or "").strip()
    raw = re.sub(r"\s*\(\s*optional\s*\)\s*", "", raw, flags=re.I)
    raw = re.sub(r"\[\s*dropdown\s*\]", "", raw, flags=re.I)
    raw = re.sub(r"\(\s*dropdown\s*\)", "", raw, flags=re.I)
    raw = re.sub(r"\(\s*select\s*\)", "", raw, flags=re.I)
    raw = raw.replace("▼", "")
    raw = raw.replace("*", "").strip()
    if not raw:
        return None
    if raw in _ELIGIBILITY_KEYS:
        return raw
    lower_key = _LOWER_KEY_MAP.get(raw.lower())
    if lower_key:
        return lower_key
    return _HEADER_ALIASES.get(_normalize_header(raw))


def build_template_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scenarios"

    export_headers = [_SCENARIO_NAME_FIELD] + _ELIGIBILITY_KEYS
    display_headers = [_SCENARIO_NAME_FIELD]
    for field in _ELIGIBILITY_KEYS:
        label = field
        if field in _OPTIONAL_FIELDS:
            label += " *"
        if field in _DROPDOWN_OPTIONS:
            label += " ▼"
        display_headers.append(label)
    ws.append(display_headers)
    sample_values = {
        _SCENARIO_NAME_FIELD: "Scenario 1",
        "occupancy": "Investment Property",
        "loanPurpose": "Purchase",
        "state": "TX",
        "valueSalesPrice": "500000",
        # loanAmount intentionally blank; formula calculates it from valueSalesPrice + ltv.
        "ltv": "70",
        "estimatedDti": "43",
        "documentationType": "Full Documentation",
        "prepaymentTerms": "No",
        "propertyType": "single_family",
        "citizenship": "US Citizen",
        "decisionCreditScore": "720",
        "paymentHistory": "0x30x12",
        "firstTimeInvestor": "No",
        "establishedPrimaryRes": "Yes",
        "isSecondLien": "No",
        "stateCounty": "Harris",
        "stateCity": "Houston",
        "stateZipCode": "77001",
        "loanTerm": "30",
        "interestOnlyPref": "No Preference",
        "rateTypePref": "No Preference",
        "lienPosition": "first_lien_only",
        "primaryLoanPurpose": "Purchase",
        "hiLavaZone": "No",
        "isRuralProperty": "No",
        "ofacSanctioned": "No",
        "hasUsCredit": "Yes",
        "investmentIncomePath": "income",
        "prepayStepdown": "No",
        "listingSeasoning": "No",
        "powerOfAttorney": "No",
        "nonArmsLength": "No",
    }
    ws.append([sample_values.get(h, "") for h in export_headers])

    header_fill = PatternFill(fill_type="solid", start_color="E2E8F0", end_color="E2E8F0")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(export_headers))}1000"

    # Auto loan amount formula: loanAmount = valueSalesPrice * ltv / 100
    # Applied for the first 1000 scenario rows.
    col_idx = {field: i + 1 for i, field in enumerate(export_headers)}
    val_col = get_column_letter(col_idx["valueSalesPrice"])
    loan_col = get_column_letter(col_idx["loanAmount"])
    ltv_col = get_column_letter(col_idx["ltv"])
    for row in range(2, 1001):
        ws[f"{loan_col}{row}"] = (
            f'=IF(AND({val_col}{row}<>"",{ltv_col}{row}<>""),ROUND({val_col}{row}*{ltv_col}{row}/100,0),"")'
        )

    # Dropdown validations for enum-like fields.
    for field, options in _DROPDOWN_OPTIONS.items():
        idx = col_idx.get(field)
        if not idx:
            continue
        col = get_column_letter(idx)
        options_csv = ",".join(options)
        dv = DataValidation(
            type="list",
            formula1=f'"{options_csv}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.showInputMessage = True
        dv.promptTitle = "Dropdown available"
        dv.prompt = "Select one value from the list."
        dv.error = "Please select a value from the dropdown."
        dv.errorTitle = "Invalid value"
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1000")

    # Sensible width for easier editing.
    for i in range(1, len(export_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


@dataclass
class _ScenarioRun:
    scenario_no: int
    scenario_name: str
    form_payload: dict[str, str]
    total_programs_evaluated: int
    input_rows: list[tuple[str, str]]
    eligible_rows: list[tuple[str, str, str]]
    loanpass_rows: list[tuple[str, str]]
    failed_rows: list[tuple[str, str, str]]
    matched_programs_contract: list[dict[str, Any]]
    rejected_programs_contract: list[dict[str, Any]]


def _parse_excel_rows(
    xlsx_bytes: bytes,
) -> list[tuple[int, str, dict[str, str], list[tuple[str, str]]]]:
    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    if not data:
        raise ValueError("Excel file is empty.")

    headers = [_cell_to_text(h) for h in data[0]]
    if not any(h.strip() for h in headers):
        raise ValueError("Header row is empty.")

    idx_to_field: dict[int, str] = {}
    scenario_col_idx: int | None = None
    for idx, head in enumerate(headers):
        if _normalize_header(head) == _normalize_header(_SCENARIO_NAME_FIELD):
            scenario_col_idx = idx
            continue
        field = _header_to_field(head)
        if field:
            idx_to_field[idx] = field

    if scenario_col_idx is None:
        raise ValueError("Template must include first column 'scenarioName'.")

    if not idx_to_field:
        raise ValueError(
            "No recognized columns found. Use /api/batch/template headers or eligibility field keys."
        )

    scenarios: list[tuple[int, str, dict[str, str], list[tuple[str, str]]]] = []
    scenario_no = 0
    for row in data[1:]:
        values = [_cell_to_text(v) for v in row]
        if not any(v.strip() for v in values):
            continue
        scenario_name = ""
        if scenario_col_idx < len(values):
            scenario_name = values[scenario_col_idx].strip()
        # Accept only rows where scenario name is explicitly provided.
        if not scenario_name:
            continue
        scenario_no += 1
        form_payload: dict[str, str] = {}
        input_rows: list[tuple[str, str]] = []
        for idx, val in enumerate(values):
            if idx == scenario_col_idx:
                continue
            if not val:
                continue
            header = headers[idx] if idx < len(headers) else f"Column {idx + 1}"
            input_rows.append((header, val))
            field = idx_to_field.get(idx)
            if field:
                form_payload[field] = val
        scenarios.append((scenario_no, scenario_name, form_payload, input_rows))

    if not scenarios:
        raise ValueError("No valid scenario rows found (scenarioName is required).")
    return scenarios


def _categorize_rejection(layer: str, reason: str) -> str:
    text = f"{layer} {reason}".lower()
    if "citizenship" in text or "itin" in text or "foreign national" in text:
        return "citizenship_mismatch"
    if "second-lien" in text or "lien" in text:
        return "lien_position_mismatch"
    if "dscr" in text and "income" in text:
        return "dscr_program_with_income_doc"
    if "below" in text and "loan" in text:
        return "loan_amount_below_minimum"
    if "above" in text and "loan" in text:
        return "loan_amount_above_maximum"
    if "ltv" in text:
        return "ltv_above_program_cap"
    if "cltv" in text:
        return "cltv_above_program_cap"
    if "fico" in text:
        return "fico_below_floor"
    if "property type" in text:
        return "property_type_not_allowed"
    if "occupancy" in text:
        return "occupancy_not_allowed"
    if "loan purpose" in text:
        return "loan_purpose_not_allowed"
    if "state" in text:
        return "state_ineligibility"
    if "county" in text:
        return "county_ineligibility"
    if "city" in text:
        return "city_ineligibility"
    if "zip" in text:
        return "zip_ineligibility"
    if "seasoning" in text or "credit event" in text:
        return "credit_event_seasoning_insufficient"
    if "overlay" in text:
        return "geo_overlay_disqualifies"
    return "geo_overlay_disqualifies"


def _rule_id_from_layer(layer: str) -> str:
    low = (layer or "").lower()
    if "layer 1" in low:
        return "program.gate"
    if "layer 2" in low:
        return "program.matrix"
    if "layer 3" in low:
        return "program.fthb"
    if "layer 4" in low:
        return "program.products"
    if "layer 5" in low:
        return "program.geo"
    if "layer 6" in low:
        return "program.credit"
    if "layer 7" in low:
        return "program.housing_history"
    if "layer 8" in low:
        return "program.guidelines"
    if "layer 10" in low:
        return "program.verify"
    return "program.rule"


def _run_scenario(
    scenario_no: int,
    scenario_name: str,
    form_payload: dict[str, str],
    input_rows: list[tuple[str, str]],
    loanpass_products_cache: dict[int, list[str]],
) -> _ScenarioRun:
    # Fast batch path: deterministic SQL engine + trace-enabled reject reasons.
    # This avoids the very slow full-mode RAG/Qdrant layers that can hang for minutes.
    result = find_eligible_programs(form_payload, quick=True, collect_trace=True)
    eligible = result.get("eligible") or []
    total_programs_evaluated = len((result.get("program_trace") or {}).get("programs") or []) or int(
        result.get("total_screened") or 0
    )

    eligible_rows: list[tuple[str, str, str]] = []
    loanpass_rows: list[tuple[str, str]] = []
    matched_contract: list[dict[str, Any]] = []

    for prog in eligible:
        program_name = str(prog.get("program_name_np") or prog.get("program_name") or "").strip()
        investor_name = str(prog.get("investor_name") or prog.get("investor") or "").strip()
        products = [str(p).strip() for p in (prog.get("products_matching") or prog.get("products") or []) if str(p).strip()]
        products_text = ", ".join(products) if products else "-"
        eligible_rows.append((program_name or "-", investor_name or "-", products_text))

        pid_raw = prog.get("program_id")
        try:
            pid = int(pid_raw) if pid_raw is not None else None
        except (TypeError, ValueError):
            pid = None
        if pid is None:
            continue

        product_items: list[dict[str, Any]] = []
        loanpass_pass = False
        loanpass_note: str | None = None
        try:
            product_names = loanpass_products_cache.get(pid)
            lp_products: list[dict[str, Any]] | None = None
            if product_names is None:
                lp = list_program_products(
                    form_payload,
                    program_id=pid,
                    program_name=program_name or None,
                    investor_name=investor_name or None,
                )
                lp_products = lp.get("products") or []
                names = [
                    str((p or {}).get("product_name") or "").strip()
                    for p in lp_products
                ]
                product_names = [p for p in names if p]
                loanpass_products_cache[pid] = product_names
            else:
                lp_products = None
            if product_names:
                loanpass_rows.append((program_name or "-", ", ".join(product_names)))
                loanpass_pass = True
            if lp_products:
                for item in lp_products:
                    name = str((item or {}).get("product_name") or "").strip()
                    if not name:
                        continue
                    lp_id = str((item or {}).get("loanpass_product_id") or "").strip()
                    product_items.append(
                        {
                            "product_type_id": (item or {}).get("product_type_id"),
                            "name": name,
                            "loanpass_product_ids": [lp_id] if lp_id else [],
                        }
                    )
        except Exception:
            # Keep report generation resilient: eligibility output is still useful even if
            # LoanPASS has a transient outage or missing mapping for some programs.
            loanpass_note = "LoanPASS lookup failed for this program in current run."

        if not product_items:
            # Fallback from internal eligibility product list when LoanPASS mapping isn't available.
            for name in products:
                product_items.append(
                    {
                        "product_type_id": None,
                        "name": name,
                        "loanpass_product_ids": [],
                    }
                )

        matched_item = {
            "program_id": pid,
            "program_name": str(prog.get("program_name") or ""),
            "program_name_np": program_name,
            "lender": {
                "id": None,
                "brand_name": investor_name,
                "lender_name": investor_name,
            },
            "is_dscr_program": bool(prog.get("is_dscr")),
            "eligible_products": product_items,
            "loanpass_pass": loanpass_pass,
        }
        if (not loanpass_pass) and loanpass_note:
            matched_item["loanpass_pass_note"] = loanpass_note
        elif not loanpass_pass:
            matched_item["loanpass_pass_note"] = (
                "Program not exposed via LoanPASS - eligibility from internal engine only"
            )
        matched_contract.append(matched_item)

    failed_rows: list[tuple[str, str, str]] = []
    rejected_contract: list[dict[str, Any]] = []
    trace_data = result.get("program_trace")
    if isinstance(trace_data, dict):
        rejected = rejected_programs_from_trace(trace_data)
        for item in rejected:
            pretty_reason = humanize_reject_reason(item.layer or "", item.reason or "")
            if " - " in item.program_title:
                lender_label, program_label = item.program_title.split(" - ", 1)
            elif " — " in item.program_title:
                lender_label, program_label = item.program_title.split(" — ", 1)
            else:
                lender_label, program_label = "", item.program_title
            failed_rows.append(
                (
                    item.program_title or f"Program {item.program_id}",
                    item.layer or "-",
                    pretty_reason,
                )
            )
            rejected_contract.append(
                {
                    "program_id": item.program_id,
                    "lender": lender_label,
                    "program_name_np": program_label,
                    "rejection_category": _categorize_rejection(item.layer or "", item.reason or ""),
                    "rejection_reason": pretty_reason,
                    "rule_id": _rule_id_from_layer(item.layer or ""),
                }
            )

    return _ScenarioRun(
        scenario_no=scenario_no,
        scenario_name=scenario_name,
        form_payload=form_payload,
        total_programs_evaluated=total_programs_evaluated,
        input_rows=input_rows,
        eligible_rows=eligible_rows,
        loanpass_rows=loanpass_rows,
        failed_rows=failed_rows,
        matched_programs_contract=matched_contract,
        rejected_programs_contract=rejected_contract,
    )


def _build_profile_sections(run: _ScenarioRun) -> list[ProfileSection]:
    profile_rows = [ProfileRow(label=label, value=value) for label, value in run.input_rows]
    sections: list[ProfileSection] = [
        ProfileSection(title="Scenario Inputs", rows=profile_rows),
    ]
    if run.loanpass_rows:
        lp_rows = [
            ProfileRow(label=prog, value=(products or "-"))
            for prog, products in run.loanpass_rows
        ]
        sections.append(ProfileSection(title="LoanPASS Passed Programs", rows=lp_rows))
    return sections


def _build_program_items(run: _ScenarioRun) -> list[ScenarioPdfProgramItem]:
    items: list[ScenarioPdfProgramItem] = []
    for program_name, investor_name, products in run.eligible_rows:
        title = program_name
        if investor_name and investor_name != "-":
            title = f"{program_name} ({investor_name})"
        items.append(
            ScenarioPdfProgramItem(
                program_title=title,
                investor_name=investor_name if investor_name != "-" else "",
                products_display=products if products != "-" else "",
            )
        )
    return items


def _build_rejected_items(run: _ScenarioRun) -> list[ScenarioPdfRejectedItem]:
    items: list[ScenarioPdfRejectedItem] = []
    for idx, (program, layer, reason) in enumerate(run.failed_rows, start=1):
        items.append(
            ScenarioPdfRejectedItem(
                program_id=idx,
                program_title=program,
                layer=layer,
                reason=reason,
            )
        )
    return items


def _wrap(text: str, max_chars: int) -> list[str]:
    words = (text or "").split()
    if not words:
        return [""]
    lines: list[str] = []
    cur = words[0]
    for w in words[1:]:
        nxt = f"{cur} {w}"
        if len(nxt) <= max_chars:
            cur = nxt
        else:
            lines.append(cur)
            cur = w
    lines.append(cur)
    return lines


def _draw_table(
    page: fitz.Page,
    y: float,
    left: float,
    right: float,
    headers: list[str],
    rows: list[tuple[str, ...]],
    col_widths: list[float],
    start_idx: int = 0,
) -> tuple[float, int]:
    border = (0.8, 0.8, 0.85)
    # Use built-in Base-14 names that always exist in PyMuPDF.
    font = "helv"
    font_bold = "hebo"
    font_size = 9.5
    line_gap = 2
    page_bottom = page.rect.height - 54
    usable_w = right - left
    widths = [usable_w * w for w in col_widths]

    def row_height(values: tuple[str, ...] | list[str]) -> float:
        max_lines = 1
        for i, v in enumerate(values):
            approx_chars = max(10, int(widths[i] / (font_size * 0.55)))
            max_lines = max(max_lines, len(_wrap(v, approx_chars)))
        return 8 + max_lines * (font_size + line_gap)

    header_h = row_height(headers)
    if y + header_h > page_bottom:
        return y, start_idx

    x = left
    for idx, h in enumerate(headers):
        w = widths[idx]
        rect = fitz.Rect(x, y, x + w, y + header_h)
        page.draw_rect(rect, color=border, fill=(0.95, 0.97, 1.0), width=0.6)
        page.insert_textbox(rect + (4, 4, -4, -2), h, fontsize=font_size, fontname=font_bold, color=(0.05, 0.1, 0.2), align=0)
        x += w
    y += header_h

    row_idx = start_idx
    while row_idx < len(rows):
        row = rows[row_idx]
        rh = row_height(row)
        if y + rh > page_bottom:
            # Ensure forward progress: if a single row is too tall, render a clipped row.
            if rh > (page_bottom - y) and (page_bottom - y) > 24:
                rh = page_bottom - y - 2
            else:
                return y, row_idx
        x = left
        for idx, val in enumerate(row):
            w = widths[idx]
            rect = fitz.Rect(x, y, x + w, y + rh)
            fill = (1, 1, 1) if row_idx % 2 == 0 else (0.985, 0.985, 0.99)
            page.draw_rect(rect, color=border, fill=fill, width=0.5)
            page.insert_textbox(
                rect + (4, 4, -4, -2),
                str(val)[:1200],
                fontsize=font_size,
                fontname=font,
                color=(0.1, 0.15, 0.2),
                align=0,
            )
            x += w
        y += rh
        row_idx += 1

    return y, row_idx


def _ensure_page(doc: fitz.Document, page: fitz.Page, y: float, needed: float) -> tuple[fitz.Page, float]:
    if y + needed <= page.rect.height - 54:
        return page, y
    page = doc.new_page(width=612, height=792)
    return page, 54


def _add_section_title(doc: fitz.Document, page: fitz.Page, y: float, text: str) -> tuple[fitz.Page, float]:
    page, y = _ensure_page(doc, page, y, 28)
    page.insert_text((54, y), text, fontsize=12, fontname="hebo", color=(0.07, 0.17, 0.35))
    return page, y + 16


def _execute_batch_runs(xlsx_bytes: bytes) -> list[_ScenarioRun]:
    parsed = _parse_excel_rows(xlsx_bytes)
    loanpass_products_cache: dict[int, list[str]] = {}
    return [
        _run_scenario(num, name, payload, inputs, loanpass_products_cache)
        for num, name, payload, inputs in parsed
    ]


def _build_contract_from_runs(scenario_runs: list[_ScenarioRun]) -> dict[str, Any]:
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    contract: dict[str, Any] = {
        "endpoint": "POST /api/eligibility-batch-test",
        "description": (
            "Runs one or more loan scenarios through the eligibility engine and returns, "
            "per scenario, matched programs (with eligible products + LoanPASS product IDs) "
            "and rejected programs (with structured rejection reasons). Batch-capable: 1..N scenarios per call."
        ),
        "version": "v6.2.1",
        "request_id": f"req_{int(datetime.now(timezone.utc).timestamp())}",
        "generated_at": now,
        "evaluation_engine_version": "v6.2.1",
        "loanpass_snapshot_at": now,
        "results": [],
    }
    for run in scenario_runs:
        scenario_id = f"SCENARIO_{run.scenario_no}"
        contract["results"].append(
            {
                "scenario_id": scenario_id,
                "label": run.scenario_name,
                "summary": {
                    "total_programs_evaluated": run.total_programs_evaluated,
                    "matched_count": len(run.matched_programs_contract),
                    "rejected_count": len(run.rejected_programs_contract),
                    "loanpass_pass_count": sum(
                        1 for m in run.matched_programs_contract if bool(m.get("loanpass_pass"))
                    ),
                },
                "matched_programs": run.matched_programs_contract,
                "rejected_programs": run.rejected_programs_contract,
                "warnings": [],
                "info_needed": [],
            }
        )
    return contract


def build_batch_outputs(xlsx_bytes: bytes) -> tuple[bytes, bytes]:
    scenario_runs = _execute_batch_runs(xlsx_bytes)
    merged = fitz.open()
    for run in scenario_runs:
        req = ScenarioPdfRequest(
            profile_sections=_build_profile_sections(run),
            programs=_build_program_items(run),
            rejected_programs=_build_rejected_items(run),
            scenario_description=run.scenario_name or f"Scenario {run.scenario_no}",
            form_fields=None,
        )
        one_pdf = generate_scenario_pdf_bytes(req)
        src = fitz.open("pdf", one_pdf)
        merged.insert_pdf(src)
        src.close()

    out = BytesIO()
    merged.save(out)
    merged.close()
    pdf_bytes = out.getvalue()
    json_bytes = (json.dumps(_build_contract_from_runs(scenario_runs), indent=2, ensure_ascii=False) + "\n").encode("utf-8")
    return pdf_bytes, json_bytes


def build_batch_contract_json_bytes(xlsx_bytes: bytes) -> bytes:
    _, json_bytes = build_batch_outputs(xlsx_bytes)
    return json_bytes


def build_batch_report_pdf_bytes(xlsx_bytes: bytes) -> bytes:
    pdf_bytes, _ = build_batch_outputs(xlsx_bytes)
    return pdf_bytes
